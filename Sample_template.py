from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side

# Define border style
thin_border = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000")
)

def create_sample_template():
    """Create a sample template file for JTL report generation."""
    try:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Sample Template"

        # Header
        headers = [
            "Type",
            "Module",
            "Scenario",
            "Pages",
            "Yes/No",
            "Target TPH"
        ]
        worksheet.append(headers)

        # Header styling
        dark_blue_fill = PatternFill(
            fill_type="solid",
            fgColor="FF203764"
        )
        white_bold_font = Font(
            color="FFFFFF",
            bold=True
        )

        for cell in worksheet[1]:
            cell.fill = dark_blue_fill
            cell.font = white_bold_font

        # Data
        rows = [
            ["UI", "", "Common Pages", "AB101_Enter Link in browser", "No", ""],
            ["UI", "", "Common Pages", "AB102_Login", "No", ""],
            ["UI", "", "Common Pages", "AB119_Logout", "No", ""],
            ["UI", "Lead", "AB Lead Creation", "AB103_Click on Gold Loan", "No", ""],
            ["UI", "Lead", "AB Lead Creation", "AB104_Click on medium fetch", "No", ""],
            ["UI", "Lead", "AB Lead Creation","AB105_Click on Profile Image Attachment upload file", "Yes", ""],
            ["UI", "Lead", "AB Lead Creation", "AB106_Search Document Type", "No", ""],
            ["UI", "Lead", "AB Lead Creation", "AB107_Select ID proof and ok", "No", ""],
            ["UI", "Lead", "AB Lead Creation", "AB108_ID Proof Attachment Upload File", "Yes", ""],
            ["UI", "Lead", "AB Lead Creation", "AB109_Click on Save and proceed", "Yes", 8890],
            ["UI", "Lead", "AB Lead Creation", "AB110_Click on start CKYC Verification", "No", ""],
            ["UI", "Lead", "AB Lead Creation", "AB111_Click OK", "No", ""],
            ["UI", "Lead", "AB Lead Creation", "AB112_Click On Generate KYC Form", "Yes", ""],
            ["UI", "Lead", "AB Lead Creation", "AB113_Click OK", "No", ""],
            ["UI", "Lead", "AB Lead Creation", "AB114_Click on KYC Form Verification", "No", ""],
            ["UI", "Lead", "AB Lead Creation", "AB115_Upload Document Attachment", "Yes", ""],
            ["UI", "Lead", "AB Lead Creation", "AB116_Click save", "No", ""],
            ["UI", "Lead", "AB Lead Creation", "AB117_Click Update", "No", ""],
            ["UI", "Lead", "AB Lead Creation", "AB118_Click OK", "Yes", ""],
            ["UI", "Lead", "AC Lead Search", "AC103_Click search", "Yes", ""],
            ["API", "API Lead", "CA_Gold Loan Lead  Creation _NC", "API_CA_Gold Loan Lead  Creation _NC", "Yes", 41371],
            ["API", "API Lead", "CC_Gold Loan Lead  Creation_EC", "API_CC_Gold Loan Lead  Creation_EC", "Yes", 16675],

            # Note
            ["", "", "", "", "", ""],
            ["Note:","1. Transaction Name must be same for all the scenarios which have same common pages.", "", "", "", ""],
            ["", "2. All API Scripts must Start with API (Before Naming Convention) (like API_BA103, API_CA102)", "", "", "", ""],
            ["", "3. Type, Module, and Scenario columns are copied from the template as it is using the transaction name.", "", "", "", ""],
            ["", "4. Yes/No Column is to decide which transaction to be included in PasFail Sheet ", "", "", "", ""],
            ["", "", "", "", "", ""], 
        ]

        for row in rows:
            worksheet.append(row)

        # Apply border to all cells in the sheet
        for row in worksheet.iter_rows(
            min_row=1,
            max_row=worksheet.max_row,
            min_col=1,
            max_col=6
        ):
            for cell in row:
                cell.border = thin_border

        # Enable multiple lines inside the cell
        for row_label in ['A25','B25','B26','B27','B28']:
            worksheet[row_label].font = Font(bold=True)

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()
    except Exception as e:
        print(f"ERROR creating sample template: {e}")
        raise ValueError(f"Failed to create sample template: {str(e)}")
