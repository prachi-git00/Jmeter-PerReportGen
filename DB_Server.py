from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')
)
bold = Font(name="Cambria", bold=True)

def cm_to_pixels(cm):
    return cm * 36.8

def DB_Server_Utilization(input_file_path, wb):

    #create sheet for DB Server Utilization
    currsheet = wb.create_sheet(title='DB Server Utilization', index=4)
    currsheet.sheet_view.showGridLines = False
    currsheet.column_dimensions['B'].width = 25
    cell = currsheet.cell(row=2, column=2)
    cell.value = "DB Server Utilization"
    cell.font = bold
    cell.alignment = Alignment(wrap_text=True)
    
    if input_file_path is None:
        return

    input_paths = [input_file_path] if isinstance(input_file_path, str) else input_file_path
    matching_sheets = []
    for input_path in input_paths:
        input_wb = load_workbook(input_path)
        sheet_keywords = ('db', 'db count', 'database')
        matching_sheets.extend(
            input_wb[sheet_name]
            for sheet_name in input_wb.sheetnames
            if any(keyword in sheet_name.casefold() for keyword in sheet_keywords)
        )

    if not matching_sheets:
        return

    # #create sheet for DB Server Utilization
    # currsheet = wb.create_sheet(title='DB Server Utilization', index=4)
    # currsheet.sheet_view.showGridLines = False
    # currsheet.column_dimensions['B'].width = 25
    # cell = currsheet.cell(row=2, column=2)
    # cell.value = "DB Server Utilization"
    # cell.font = bold
    # cell.alignment = Alignment(wrap_text=True)

    print("Using sheets in DB Server Utilization:", ", ".join(sheet.title for sheet in matching_sheets))

    # Start placing images from row 3 (same as your data)
    img_row = 4
    img_col = 2   # choose column (e.g., column E)

    for input_sheet in matching_sheets:
        for image in input_sheet._images:
            try:
                new_img = Image(image.ref)
            except:
                new_img = Image(image._data())

            # Set fixed size (IMPORTANT)
            new_img.width = cm_to_pixels(24)
            new_img.height = cm_to_pixels(12)

            # Place image in a clean grid (no overlap)
            cell_position = f"{get_column_letter(img_col)}{img_row}"
            currsheet.add_image(new_img, cell_position)

            # Move to next row → prevents overlap
            img_row += 26