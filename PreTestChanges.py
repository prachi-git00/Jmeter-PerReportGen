from datetime import datetime

from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from db_manager import get_pretest_changes, _parse_change_date

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')
)
bold = Font(name="Cambria", bold=True)
white_bold = Font(name="Cambria", color="FFFFFFFF", bold=True)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
dark_header = PatternFill(start_color="FF222B35", end_color="FF222B35", fill_type="solid")

start_row = 2
start_col = 2

def PreTestChanges(wb, project_name=None):
    """Add Pretest Changes sheet to workbook with exception handling."""
    try:
        currsheet = wb.create_sheet(title='Pre Test changes', index=5)
        currsheet.sheet_view.showGridLines = False

        currsheet.column_dimensions['E'].width = 70
        currsheet.column_dimensions['F'].width = 30

        headers = ['Sr. No.','Run ID','Tier','Changes Done','Date']

        for i, header in enumerate(headers):
            cell = currsheet.cell(row=start_row, column=start_col + i, value=header)
            cell.font = white_bold
            cell.border = thin_border
            cell.fill = dark_header
            cell.alignment = center
        
        # Fetch pretest changes from database
        if project_name:
            try:
                changes = get_pretest_changes(project_name)
                changes = sorted(changes, key=lambda row: _parse_change_date(row[3]))
                
                # Populate data rows
                for idx, (run_id, tier, changes_text, change_date) in enumerate(changes, 1):
                    try:
                        row = start_row + idx
                        
                        # Sr. No.
                        cell = currsheet.cell(row=row, column=start_col, value=idx)
                        cell.border = thin_border
                        cell.alignment = center
                        
                        # Run ID
                        cell = currsheet.cell(row=row, column=start_col + 1, value=run_id)
                        cell.border = thin_border
                        cell.alignment = center

                        # Tier
                        cell = currsheet.cell(row=row, column=start_col + 2, value=tier)
                        cell.border = thin_border
                        cell.alignment = center
                        
                        # Changes Done
                        cell = currsheet.cell(row=row, column=start_col + 3, value=changes_text)
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                        
                        # Date
                        try:
                            parsed_date = datetime.strptime(str(change_date), "%Y-%m-%d").date()
                        except ValueError:
                            try:
                                parsed_date = datetime.strptime(str(change_date), "%Y-%m-%d %H:%M:%S").date()
                            except ValueError:
                                parsed_date = None

                        cell = currsheet.cell(row=row, column=start_col + 4, value=parsed_date)
                        if parsed_date is not None:
                            cell.number_format = 'dd-mm-yyyy'
                        cell.border = thin_border
                        cell.alignment = center
                    except Exception as e:
                        print(f"❌ Error adding pretest change row {idx}: {e}")
                        continue
            except Exception as e:
                print(f"❌ Error retrieving pretest changes for project '{project_name}': {e}")
    except Exception as e:
        print(f"❌ Error creating Pretest Changes sheet: {e}")
