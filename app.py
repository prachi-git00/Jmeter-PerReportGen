import os
import tempfile
from io import BytesIO
from datetime import date, datetime, timedelta

# import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
# from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from Sample_template import create_sample_template

from Cal_TPH_RespTime_jtl import (
    calculate_average_throughput,
    generate_jtl_summary_sheets,
    load_jtl_data,
    load_template,
)
from DB_Server import DB_Server_Utilization
from Errors import Errors
from ExeSummary import ExecutiveSummary
from JTLPreview import show_jtl_preview
from PodService import podsServiceUtilization
from PreTestChanges import PreTestChanges

from db_manager import (
    add_pretest_change,
    create_project,
    delete_project,
    delete_pretest_change,
    get_all_projects,
    get_pretest_changes_filtered,
    init_database,
    update_pretest_change,
)


st.set_page_config(page_title="JMeter Performance Report Generator", page_icon="📊", layout="wide")
st.title("📊 JMeter Performance Report Generator")

try:
    template_data_bytes = create_sample_template()
    st.download_button(
        label="Download Sample Template",
        data=template_data_bytes,
        file_name="Sample_Template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
except Exception as e:
    st.error(f"❌ Error creating sample template: {str(e)}")

@st.cache_resource
def initialize_database():
    try:
        init_database()
        return True
    except Exception as e:
        print(f"❌ Database initialization error: {e}")
        st.error(f"❌ Could not initialize database: {str(e)}")
        return False

try:
    db_initialized = initialize_database()
    if not db_initialized:
        st.stop()
except Exception as e:
    st.error(f"❌ Application initialization failed: {str(e)}")
    st.stop()

st.sidebar.header("📁 Project Management")

with st.sidebar.expander("➕ Create New Project"):
    new_project_name = st.text_input("Project Name")
    if st.button("Create Project"):
        if new_project_name.strip():
            if create_project(new_project_name):
                st.success(f"✅ Project '{new_project_name}' created!")
            else:
                st.error(f"❌ Project '{new_project_name}' already exists!")
        else:
            st.error("Please enter a project name")

all_projects = get_all_projects()
if all_projects:
    selected_project = st.sidebar.selectbox("Select Project", all_projects, key="project_select")
else:
    selected_project = None
    st.sidebar.info("No projects created yet. Create one first!")

if selected_project:
    with st.sidebar.expander("🗑️ Delete Project"):
        if st.button("Delete Selected Project", key="delete_project_button"):
            st.session_state.delete_project_confirmation = selected_project

        if st.session_state.get("delete_project_confirmation") == selected_project:
            st.warning(f"Delete '{selected_project}' and all its pretest changes?")
            confirm_delete = st.checkbox("Confirm project deletion", key="confirm_delete_project")
            if st.button("Confirm Delete", key="confirm_delete_project_button"):
                if confirm_delete and delete_project(selected_project):
                    del st.session_state.delete_project_confirmation
                    st.success(f"✅ Deleted project '{selected_project}'")
                    st.rerun()
                else:
                    st.error("Please confirm the project deletion first")

if 'edit_run_id' not in st.session_state:
    st.session_state.edit_run_id = None
    st.session_state.edit_tier = ''
    st.session_state.edit_changes = ''
    st.session_state.edit_date = None

if selected_project:
    st.sidebar.header("📝 Pretest Changes")

    with st.sidebar.expander("➕ Add New Pretest Change"):
        run_id = st.text_input("Run ID", key="run_id_input")
        tier = st.text_input("Tier", key="tier_input")
        changes_text = st.text_area("Changes Done", key="changes_input", height=100)
        change_date = st.date_input("Date", key="date_input")

        if st.button("Save Pretest Change"):
            if run_id.strip() and tier.strip() and changes_text.strip():
                success, is_duplicate = add_pretest_change(selected_project, run_id, tier, changes_text, str(change_date))
                if success:
                    if is_duplicate:
                        st.warning(f"⚠️ Run ID '{run_id}' already exists in this project. A duplicate record has been added.")
                    else:
                        st.success(f"✅ Pretest change for Run ID '{run_id}' added!")
                else:
                    st.error("Failed to add pretest change")
            else:
                st.error("Please fill in all fields")

    with st.sidebar.expander("🔎 Search / Filter Pretest Changes", expanded=True):
        filter_run_id = st.text_input("Filter by Run ID", key="filter_run_id")
        filter_tier = st.text_input("Filter by Tier", key="filter_tier")
        date_filter = st.checkbox("Filter by date range", key="filter_by_date")
        if date_filter:
            col1, col2 = st.columns(2)
            with col1:
                start_date = st.date_input("From", key="filter_start_date")
            with col2:
                end_date = st.date_input("To", key="filter_end_date")
        else:
            start_date = None
            end_date = None

    with st.sidebar.expander("📋 View Pretest Changes"):
        changes_list = get_pretest_changes_filtered(
            selected_project,
            run_id=filter_run_id,
            tier=filter_tier,
            start_date=str(start_date) if start_date else None,
            end_date=str(end_date) if end_date else None,
        )

        if changes_list:
            st.markdown(f"**Showing {len(changes_list)} records**")
            for idx, (run_id, tier, changes_text, change_date) in enumerate(changes_list, 1):
                display_date = ""
                try:
                    display_date = datetime.strptime(str(change_date), "%Y-%m-%d").strftime("%d-%m-%Y")
                except ValueError:
                    try:
                        display_date = datetime.strptime(str(change_date), "%Y-%m-%d %H:%M:%S").strftime("%d-%m-%Y")
                    except ValueError:
                        display_date = str(change_date)

                st.write(f"**{idx}. Run ID: {run_id}** | Tier: {tier} ({display_date})")
                st.caption(changes_text)
                cols = st.columns([3, 1])
                with cols[0]:
                    if st.button("Edit", key=f"edit_{idx}_{run_id}"):
                        st.session_state.edit_run_id = run_id
                        st.session_state.edit_tier = tier
                        st.session_state.edit_changes = changes_text
                        try:
                            st.session_state.edit_date = datetime.strptime(str(change_date), "%Y-%m-%d").date()
                        except Exception:
                            try:
                                st.session_state.edit_date = datetime.strptime(str(change_date), "%Y-%m-%d %H:%M:%S").date()
                            except Exception:
                                st.session_state.edit_date = date.today()
                with cols[1]:
                    if st.button("Delete", key=f"delete_{idx}_{run_id}"):
                        if delete_pretest_change(selected_project, run_id):
                            st.success(f"✅ Deleted Run ID '{run_id}'")
                st.divider()
        else:
            st.info("No pretest changes match the current filters")

    if st.session_state.edit_run_id:
        edit_date_obj = None
        if isinstance(st.session_state.edit_date, str):
            try:
                edit_date_obj = datetime.strptime(st.session_state.edit_date, "%Y-%m-%d").date()
            except Exception:
                edit_date_obj = None
        elif isinstance(st.session_state.edit_date, datetime):
            edit_date_obj = st.session_state.edit_date.date()
        elif isinstance(st.session_state.edit_date, date):
            edit_date_obj = st.session_state.edit_date

        if edit_date_obj is None:
            edit_date_obj = date.today()

        with st.sidebar.expander(f"✏️ Edit Pretest Change: {st.session_state.edit_run_id}", expanded=True):
            edit_tier = st.text_input("Tier", value=st.session_state.edit_tier, key="edit_tier_input")
            edit_changes = st.text_area("Changes Done", value=st.session_state.edit_changes, key="edit_changes_input", height=100)
            edit_date = st.date_input("Date", value=edit_date_obj, key="edit_date_input")

            if st.button("Update Pretest Change"):
                if edit_tier.strip() and edit_changes.strip():
                    if update_pretest_change(selected_project, st.session_state.edit_run_id, edit_tier, edit_changes, str(edit_date)):
                        st.success(f"✅ Updated Run ID '{st.session_state.edit_run_id}'")
                        st.session_state.edit_run_id = None
                        st.session_state.edit_tier = ''
                        st.session_state.edit_changes = ''
                        st.session_state.edit_date = None
                    else:
                        st.error("Failed to update pretest change")
                else:
                    st.error("Tier and Changes cannot be empty")

st.subheader("📈 JTL Performance Analysis")

jtl_files = st.file_uploader(
    "Upload JTL Files",
    type=["jtl", "csv"],
    accept_multiple_files=True,
    key="jtl_performance_files",
)

template_file = st.file_uploader(
    "Upload Transaction Template",
    type=["xlsx", "xls"],
    key="transaction_template_file",
)

jtl_col1, jtl_col2 = st.columns(2)
with jtl_col1:
    jtl_from_date = st.date_input("JTL From Date", value=date.today(), key="jtl_from_date")
    jtl_from_time = st.time_input(
        "JTL From Time",
        value=datetime.now().time(),
        key="jtl_from_time",
        step=timedelta(minutes=1),
    )

with jtl_col2:
    jtl_to_date = st.date_input("JTL To Date", value=date.today(), key="jtl_to_date")
    jtl_to_time = st.time_input(
        "JTL To Time",
        value=datetime.now().time(),
        key="jtl_to_time",
        step=timedelta(minutes=1),
    )

Pods_and_DB_Count = st.file_uploader(
    "Upload Pods Excel / DB Count Screenshot",
    type=["xlsx", "xls"],
    accept_multiple_files=True,
    key="pods_db_file",
)

RunID = st.text_input("Enter Run ID")

@st.cache_data(show_spinner=False)
def get_cached_jtl_data(file_payloads):
    return load_jtl_data([BytesIO(payload) for payload in file_payloads])


@st.cache_data(show_spinner=False)
def get_cached_template(template_payload):
    return load_template(BytesIO(template_payload))


jtl_data = None
if jtl_files:
    try:
        jtl_data = get_cached_jtl_data(tuple(file.getvalue() for file in jtl_files))
    except Exception:
        pass

template_data = None
if template_file:
    try:
        template_data = get_cached_template(template_file.getvalue())
    except Exception:
        pass

if jtl_files:
    with st.expander("📊 Preview JTL Summary", expanded=True):
        show_jtl_preview(
            jtl_files=jtl_files,
            from_date=jtl_from_date,
            from_time=jtl_from_time,
            to_date=jtl_to_date,
            to_time=jtl_to_time,
            template_file=template_file,
            data=jtl_data,
            template=template_data,
        )

if st.button("Generate Report"):
    if not selected_project:
        st.error("⚠️ Please select or create a project first!")
        st.stop()

    if not jtl_files:
        st.error("Please upload at least one JTL file.")
        st.stop()

    if not template_file:
        st.error("Please upload a transaction template file.")
        st.stop()

    from_dt = datetime.combine(jtl_from_date, jtl_from_time)
    to_dt = datetime.combine(jtl_to_date, jtl_to_time)
    if from_dt >= to_dt:
        st.error("JTL From Date/Time cannot be greater than or equal to JTL To Date/Time.")
        st.stop()

    with tempfile.TemporaryDirectory() as temp_dir:
        wb = Workbook()
        wb.remove(wb.active)

        try:
            # Generate JTL summary sheets with exception handling
            try:
                result = generate_jtl_summary_sheets(
                    wb,
                    jtl_files,
                    template_file,
                    from_date=from_dt,
                    to_date=to_dt,
                    pass_only=True,
                    data=jtl_data,
                )
                passfail_total_rows, grand_total_row, response_total_row = result
            except Exception as e:
                st.error(f"❌ Error generating JTL summary sheets: {str(e)}")
                raise

            # Calculate throughput with exception handling
            try:
                avg_throughput = calculate_average_throughput(
                    jtl_files,
                    from_date=from_dt,
                    to_date=to_dt,
                    template=template_data,
                    data=jtl_data,
                )
            except Exception as e:
                st.warning(f"⚠️ Could not calculate average throughput: {str(e)}")
                avg_throughput = None

            # Process Pods and DB files with exception handling
            if Pods_and_DB_Count:
                try:
                    pods_db_paths = []
                    for uploaded_file in Pods_and_DB_Count:
                        try:
                            pods_db_path = os.path.join(temp_dir, uploaded_file.name)
                            with open(pods_db_path, "wb") as f:
                                f.write(uploaded_file.getbuffer())
                            pods_db_paths.append(pods_db_path)
                        except Exception as e:
                            st.warning(f"⚠️ Could not save file {uploaded_file.name}: {str(e)}")
                            continue

                    if pods_db_paths:
                        try:
                            podsServiceUtilization(pods_db_paths, wb)
                        except Exception as e:
                            st.warning(f"⚠️ Could not process Pod Service Utilization: {str(e)}")
                        
                        try:
                            DB_Server_Utilization(pods_db_paths, wb)
                        except Exception as e:
                            st.warning(f"⚠️ Could not process DB Server Utilization: {str(e)}")
                except Exception as e:
                    st.warning(f"⚠️ Error processing infrastructure files: {str(e)}")

            # Add error sheet with exception handling
            try:
                Errors(wb)
            except Exception as e:
                st.warning(f"⚠️ Could not generate Errors sheet: {str(e)}")

            # Add pretest changes sheet with exception handling
            try:
                PreTestChanges(wb, selected_project)
            except Exception as e:
                st.warning(f"⚠️ Could not add Pretest Changes sheet: {str(e)}")

            # Generate executive summary with exception handling
            try:
                ExecutiveSummary(wb, RunID, passfail_total_rows, grand_total_row, response_total_row, avg_throughput)
            except Exception as e:
                st.warning(f"⚠️ Could not generate Executive Summary sheet: {str(e)}")

            # Save workbook with exception handling
            try:
                file_date = jtl_from_date.strftime("%d%m%Y")
                report_file_name = f"{selected_project}_R{RunID}_LoadTestReport_{file_date}.xlsx"
                output_path = os.path.join(temp_dir, report_file_name)
                wb.save(output_path)
            except Exception as e:
                st.error(f"❌ Error saving workbook: {str(e)}")
                raise

            st.success("✅ JMeter Performance Report generated successfully!")
            with open(output_path, "rb") as f:
                st.markdown(
                    """
                    <style>
                    div[data-testid="stDownloadButton"] button {
                        color: #000000;
                        background-color: #ffffff;
                        font-weight: 1000;
                        transition: background-color 0.15s ease;
                    }
                    div[data-testid="stDownloadButton"] button:hover {
                        color: #000000;
                        background-color: #E0E0E0;
                    }
                    </style>
                    """,
                    unsafe_allow_html=True,
                )
                st.download_button(
                    label="📥 Download JMeter Report",
                    data=f.read(),
                    file_name=report_file_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="secondary",
                )
        except Exception as e:
            st.error(f"❌ Report generation failed: {str(e)}")
            import traceback
            st.error(f"Debug traceback: {traceback.format_exc()}")
