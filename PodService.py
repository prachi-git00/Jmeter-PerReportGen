from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.drawing.image import Image


center = Alignment(horizontal="center",vertical="center",wrap_text=True)
bold = Font(name="Cambria",bold=True)
thin_border = Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style="thin"))

def cm_to_pixels(cm):
    # Convert centimeters to pixels.
    return int(cm * 37.795)

# Main Function
def pods_service_utilization(input_file_path, wb):

    currsheet = wb.create_sheet(
        title='App Server Utilization',
        index=3
    )

    currsheet.sheet_view.showGridLines = False

    cell = currsheet.cell(row=2, column=2)
    cell.value = "App Services Pod Wise Utilization"
    cell.font = bold
    cell.alignment = Alignment(wrap_text=True)

    # Column widths
    currsheet.column_dimensions['B'].width = 25
    currsheet.column_dimensions['I'].width = 25

    if input_file_path is None:
        return

    input_paths = (
        [input_file_path]
        if isinstance(input_file_path, str)
        else input_file_path
    )

    # Find matching sheets
    matching_sheets = []

    for input_path in input_paths:
        input_wb = load_workbook(input_path)
        sheet_keywords = (
            'pods',
            'infra',
            'cpu utilization',
            'app server'
        )

        for sheet_name in input_wb.sheetnames:
            if any(keyword in sheet_name.casefold() for keyword in sheet_keywords):
                matching_sheets.append(input_wb[sheet_name])

    if not matching_sheets:
        return

    # Header definition
    headers = [
        '',
        'CPU Usage',
        'CPU Requests',
        'CPU Requests%',
        'CPU Limits',
        'CPU Limits%',
        '',
        '',
        'Memory Usage',
        'Memory Requests',
        'Memory Requests%',
        'Memory Limits',
        'Memory Limits%',
        'Memory Usage (RSS)',
        'Memory Usage (Cache)',
        'Memory Usage (Swap)'
    ]

    output_row = 5

    headers_added = False

    for input_sheet in matching_sheets:

        print(
            f"Processing sheet: {input_sheet.title}"
        )

        has_data = False

        for row in input_sheet.iter_rows(
            min_row=4,
            values_only=True
        ):
            if any(
                value is not None
                and str(value).strip() != ''
                for value in row
            ):
                has_data = True
                break

     # Get images from the input sheet
        images = getattr(input_sheet,'_images',[])

        print(f"  Table data: {has_data}")
        print(f"  Images found: {len(images)}")

        # CASE 1:
        # Input sheet contains TABLE DATA
        if has_data:
            if not headers_added:

                # Title
                title_cell = currsheet.cell(row=2,column=2)
                title_cell.value = ("App Services Pod Wise Utilization")

                title_cell.font = bold
                title_cell.alignment = Alignment(wrap_text=True)

                # Headers
                for i, header in enumerate(headers):
                    cell = currsheet.cell(
                        row=4,
                        column=2 + i,
                        value=header
                    )
                    cell.font = bold
                    cell.border = thin_border
                    cell.alignment = center

                headers_added = True

            # Copy table data
 
            rows_written = 0
            for i, row in enumerate(
                input_sheet.iter_rows(min_row=4,values_only=True)
            ):
                # Maximum 200 rows
                if i >= 200:
                    break

                # Skip completely empty rows
                if not any(
                    value is not None
                    and str(value).strip() != ''
                    for value in row
                ):
                    continue

                # Copy each cell
                for col_index, value in enumerate(row,start=1):
                    cell = currsheet.cell(row=output_row,column=col_index,value=value)
                    cell.font = Font(name="Cambria")

                    # Columns B and I
                    # are left aligned
                    if col_index in [2, 9]:
                        cell.alignment = Alignment(vertical="center",wrap_text=True)

                    else:
                        cell.alignment = Alignment(vertical="center",horizontal="center",wrap_text=True)

                    # Don't put border on column A
                    if col_index != 1:
                        cell.border = thin_border

                # Row height
                currsheet.row_dimensions[output_row].height = 28

                # Percentage formatting
                currsheet[f'E{output_row}'].number_format = '0.00%'
                currsheet[f'G{output_row}'].number_format = '0.00%'
                currsheet[f'L{output_row}'].number_format = '0.00%'
                currsheet[f'N{output_row}'].number_format = '0.00%'

                # Move to next row
                output_row += 1
                rows_written += 1

            print(
                f"  Rows copied: {rows_written}"
            )

        # CASE 2:
        # Input sheet contains IMAGES
        if images:
            # Images are placed after the table.If there was no table, they are placed at the current output row.
            
            image_row = output_row + 2
            for image_index, image in enumerate(images,start=1):
                # Create a new image object
                try:
                    new_img = Image(mage.ref)
                except Exception:
                    new_img = Image(image._data())

                # Resize image
                new_img.width = cm_to_pixels(24)
                new_img.height = cm_to_pixels(12)

                # Image position

                image_position = (f"B{image_row}")
                currsheet.add_image(new_img,image_position)
                print(
                    f"  Image {image_index} "
                    f"placed at {image_position}"
                )

                # Move down for next image
                # 26 rows gives enough vertical space for
                # a 12 cm image.
                image_row += 26

            output_row = image_row + 1

    # Column widths
    for col in [
        'C',
        'D',
        'E',
        'F',
        'G',
        'H',
        'J',
        'K',
        'L',
        'M',
        'N',
        'O',
        'P',
        'Q'
    ]:
        currsheet.column_dimensions[col].width = 10

    print("App Server Utilization sheet created successfully.")

