from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')
)
bold = Font(name="Cambria", bold=True)

def Errors(wb):
    currsheet = wb.create_sheet(title='Errors', index=6)
    currsheet.sheet_view.showGridLines = False
    currsheet.column_dimensions['B'].width = 176

    cell = currsheet.cell(row=2, column=2)
    cell.value = "Errors Observed"
    cell.border = thin_border
    cell.font = bold