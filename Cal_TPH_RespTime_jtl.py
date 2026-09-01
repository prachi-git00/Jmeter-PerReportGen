import re
import unicodedata

import pandas as pd
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

HEADER_ROW = 23
start_row = 4
start_col = 2

# ARGB colors
dark_header = PatternFill(start_color="FF4A4A98", end_color="FF4A4A98", fill_type="solid")
fill = PatternFill(start_color="FFCCCCFF", end_color="FFCCCCFF", fill_type="solid")
Header_fill = PatternFill(start_color="FF203764", end_color="FF203764", fill_type="solid")
highlight_90 = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")

white_bold = Font(name="Cambria", color="FFFFFFFF", bold=True)
bold = Font(name="Cambria", bold=True)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')
)

thin_side = Side(style='thin')


def _merge_consecutive_cells(ws, values, column, first_row):
    """Merge adjacent worksheet cells when their source values are identical."""
    if not values:
        return

    group_start = 0
    for index in range(1, len(values) + 1):
        if index == len(values) or values[index] != values[group_start]:
            if index - group_start > 1:
                ws.merge_cells(
                    start_row=first_row + group_start,
                    start_column=column,
                    end_row=first_row + index - 1,
                    end_column=column,
                )
            group_start = index


def _normalize_label(value):
    """Create a stable comparison key for labels read from Excel and JTL files."""
    normalized = unicodedata.normalize("NFKC", str(value))
    normalized = normalized.replace("\ufeff", "").replace("\u200b", "")
    return " ".join(normalized.split()).casefold()


def _load_template(template_file):
    """Read and validate the transaction template used by all report calculations."""
    if template_file is None:
        raise ValueError("Please upload a transaction template file.")
    if hasattr(template_file, "seek"):
        template_file.seek(0)

    template = pd.read_excel(template_file, header=None)

    def normalize_column_name(value):
        return re.sub(r"[^a-z0-9]", "", str(value).strip().lower())

    # Locate the actual header row when the workbook has a title or blank rows above it.
    header_row = None
    header_aliases = {
        "page": {"page", "pages", "pagename", "pagesname", "pagenames", "transaction", "transactionname", "transactionlabel"},
        "scenario": {"typemodulescenario", "typemodulescenarioname", "scenario"},
        "selection": {"yesno", "passfail", "include", "select", "selected"},
    }
    for row_index in range(min(10, len(template))):
        normalized_headers = {normalize_column_name(value) for value in template.iloc[row_index]}
        if all(aliases & normalized_headers for aliases in header_aliases.values()):
            header_row = row_index
            break

    if header_row is None:
        raise ValueError("Template must contain Page, Type Module Scenario, and Yes/No columns.")

    template.columns = template.iloc[header_row]
    template = template.iloc[header_row + 1:].reset_index(drop=True)
    column_names = {
        normalize_column_name(column): column for column in template.columns
    }
    required = {"page", "scenario", "selection"}
    missing = required.difference(
        key for key, aliases in header_aliases.items()
        if aliases & set(column_names)
    )
    if missing:
        missing_text = ", ".join(sorted(missing))
        raise ValueError(f"Template is missing required column(s): {missing_text}.")

    template = template.rename(columns={
        column_names[next(key for key in column_names if key in header_aliases["page"])]: "Page",
        column_names[next(key for key in column_names if key in header_aliases["scenario"])]: "Type Module Scenario",
        column_names[next(key for key in column_names if key in header_aliases["selection"])]: "Yes/No",
    })
    type_column = next((column for key, column in column_names.items() if key == "type"), None)
    module_column = next((column for key, column in column_names.items() if key == "module"), None)
    target_tph_column = next(
        (column for key, column in column_names.items() if key in {"targettph", "targetthroughput"}),
        None,
    )
    if type_column is not None:
        template = template.rename(columns={type_column: "Type"})
    if module_column is not None:
        template = template.rename(columns={module_column: "Module"})
    if target_tph_column is not None:
        template = template.rename(columns={target_tph_column: "Target TPH"})
    if "Type" not in template.columns:
        template["Type"] = ""
    if "Module" not in template.columns:
        template["Module"] = ""
    if "Target TPH" not in template.columns:
        template["Target TPH"] = ""
    template = template.rename(columns={"Type Module Scenario": "Scenario"})
    template["Page"] = template["Page"].astype(str).str.strip()
    template = template[template["Page"].ne("") & template["Page"].ne("nan")].copy()
    template["Yes/No"] = template["Yes/No"].astype(str).str.strip().str.lower()
    template["Type"] = template["Type"].fillna("")
    template["Module"] = template["Module"].fillna("")
    template["Scenario"] = template["Scenario"].fillna("")
    template["Target TPH"] = template["Target TPH"].fillna("")
    return template.drop_duplicates("Page", keep="first").reset_index(drop=True)


def load_template(template_file):
    """Load and validate a transaction template for reuse by report calculations."""
    return _load_template(template_file)


def load_jtl_data(jtl_files):
    """Read and normalize uploaded JTL files once for downstream calculations."""
    dfs = []
    for jtl_file in jtl_files:
        if hasattr(jtl_file, "seek"):
            jtl_file.seek(0)
        for chunk in pd.read_csv(jtl_file, chunksize=50_000, low_memory=False):
            dfs.append(chunk)

    if not dfs:
        raise ValueError("No JTL data found.")

    return _normalize_jtl_data(pd.concat(dfs, ignore_index=True))


def _common_page_keys(template):
    """Return template page keys marked as common in Module or Scenario."""
    common_mask = (
        template["Module"].astype(str).str.contains("common", case=False, na=False)
        | template["Scenario"].astype(str).str.contains("common", case=False, na=False)
    )
    return set(template.loc[common_mask, "Page"].map(_normalize_label))


def _normalize_jtl_data(data):
    """Validate and normalize the columns shared by all JTL calculations."""
    required_columns = {"success", "timeStamp", "elapsed", "label"}
    missing_columns = required_columns.difference(data.columns)
    if missing_columns:
        missing = ", ".join(sorted(missing_columns))
        raise ValueError(f"JTL file does not contain required column(s): {missing}.")

    data["label"] = data["label"].astype(str).str.strip()
    data["label_key"] = data["label"].map(_normalize_label)
    data["success_bool"] = data["success"].astype(str).str.strip().str.lower() == "true"
    data["timeStamp"] = pd.to_datetime(
        data["timeStamp"], format="%Y/%m/%d %H:%M:%S", errors="coerce"
    )
    data["elapsed"] = pd.to_numeric(data["elapsed"], errors="coerce")
    data = data.dropna(subset=["timeStamp", "elapsed"])
    data["elapsed_sec"] = data["elapsed"] / 1000
    return data


def calculate_response_jtl(
    jtl_files,
    from_date=None,
    to_date=None,
    pass_only=False,
    keywords=None,
    template=None,
    data=None,
):
    """Process JTL files and return a performance summary DataFrame."""
    if template is not None and not isinstance(template, pd.DataFrame):
        template = _load_template(template)
    merged = load_jtl_data(jtl_files) if data is None else _normalize_jtl_data(data.copy())

    if from_date is not None:
        from_date = pd.to_datetime(from_date)
        merged = merged[merged["timeStamp"] >= from_date]

    if to_date is not None:
        to_date = pd.to_datetime(to_date)
        if to_date.time() == pd.Timestamp("1900-01-01 00:00:00").time():
            to_date = to_date + pd.Timedelta(days=1)
        merged = merged[merged["timeStamp"] <= to_date]

    # Exclude internal scripting samplers from user-facing transaction summaries.
    sampler_pattern = r"beanshell\s+sampler|jsr223\s+sampler"
    merged = merged[
        ~merged["label"].fillna("").astype(str).str.contains(sampler_pattern, case=False, na=False)
    ].copy()

    if template is not None:
        template_keys = set(template["Page"].map(_normalize_label))
        merged = merged[merged["label_key"].isin(template_keys)].copy()

    if keywords:
        # Optionally restrict the summary to labels containing one of the supplied keywords.
        keyword_list = [k.strip() for k in str(keywords).split(',') if k.strip()]
        if keyword_list:
            pattern = "|".join(re.escape(k.lower()) for k in keyword_list)
            merged = merged[merged["label"].astype(str).str.lower().str.contains(pattern, regex=True, na=False)].copy()

    total_per_label = merged.groupby("label").size().rename("TotalSamples")

    if pass_only:
        # ResponseTime uses successful samples only and calculates percentile metrics per label.
        merged = merged[merged["success_bool"]].reset_index(drop=True)
        summary = (
            merged.groupby("label")
            .agg(
                Samples=("elapsed_sec", "count"),
                Min=("elapsed_sec", "min"),
                Avg=("elapsed_sec", "mean"),
                Max=("elapsed_sec", "max"),
                P90=("elapsed_sec", lambda x: x.quantile(0.90)),
                P95=("elapsed_sec", lambda x: x.quantile(0.95)),
                P99=("elapsed_sec", lambda x: x.quantile(0.99)),
            )
            .reset_index()
        )
        summary = summary.join(total_per_label, on="label")
        summary["PassPercent"] = (summary["Samples"] / summary["TotalSamples"] * 100).round(3)
        summary.drop(columns=["TotalSamples"], inplace=True)
    else:
        # PassFail uses all samples and calculates the error percentage per label.
        summary = (
            merged.groupby("label")
            .agg(
                Samples=("elapsed_sec", "count"),
                Min=("elapsed_sec", "min"),
                Avg=("elapsed_sec", "mean"),
                Max=("elapsed_sec", "max"),
                P90=("elapsed_sec", lambda x: x.quantile(0.90)),
                P95=("elapsed_sec", lambda x: x.quantile(0.95)),
                P99=("elapsed_sec", lambda x: x.quantile(0.99)),
                ErrorPercent=("success_bool", lambda x: (x.eq(False).sum() / len(x)) * 100),
            )
            .reset_index()
        )

    if template is not None:
        # Keep every template page visible, including pages with no samples in the selected window.
        template_pages = template[["Page"]].copy()
        template_pages["label_key"] = template_pages["Page"].map(_normalize_label)
        summary["label_key"] = summary["label"].map(_normalize_label)
        summary = template_pages.merge(summary, on="label_key", how="left")
        summary["label"] = summary["Page"]
        summary.drop(columns=["Page", "label_key"], inplace=True)
        metric_columns = [column for column in summary.columns if column != "label"]
        summary[metric_columns] = summary[metric_columns].fillna(0).infer_objects(copy=False)

    if not merged.empty:
        duration = (merged["timeStamp"].max() - merged["timeStamp"].min()).total_seconds()
    else:
        duration = 0

    # Calculate transaction throughput over the timestamp range represented by the filtered data.
    throughput = len(merged) / duration if duration > 0 else 0
    summary["Throughput"] = round(throughput, 2)
    return summary.round(3)


def calculate_average_throughput(jtl_files, from_date=None, to_date=None, template=None, data=None):
    """Calculate average received and sent throughput for the selected JTL samples."""
    if template is not None and not isinstance(template, pd.DataFrame):
        template = _load_template(template)
    throughput_data = load_jtl_data(jtl_files) if data is None else _normalize_jtl_data(data.copy())
    # Throughput requires byte counts, timing, success status, labels, and parent-row identification.
    required_columns = {"timeStamp", "elapsed", "bytes", "sentBytes", "success", "label"}
    missing_columns = required_columns.difference(throughput_data.columns)
    if missing_columns:
        missing = ", ".join(sorted(missing_columns))
        raise ValueError(f"JTL file does not contain required throughput column(s): {missing}.")

    # Count only successful transactions in the throughput totals.
    throughput_data = throughput_data[
        throughput_data["success"].astype(str).str.strip().str.lower() == "true"
    ]

    if from_date is not None:
        # Apply the selected time window before calculating the duration and byte rates.
        throughput_data = throughput_data[throughput_data["timeStamp"] >= pd.to_datetime(from_date)]

    if to_date is not None:
        end_date = pd.to_datetime(to_date)
        if end_date.time() == pd.Timestamp("1900-01-01 00:00:00").time():
            end_date += pd.Timedelta(days=1)
        throughput_data = throughput_data[throughput_data["timeStamp"] <= end_date]

    sampler_pattern = r"beanshell\s+sampler|jsr223\s+sampler"
    # Throughput is based on successful parent transactions, excluding setup and teardown labels.
    throughput_data = throughput_data[
        ~throughput_data["label"].fillna("").astype(str).str.contains(
            sampler_pattern, case=False, na=False
        )
    ].copy()

    throughput_data = throughput_data[
        ~throughput_data["label"].fillna("").astype(str).str.contains(
            r"login|landing|logout", case=False, na=False
        )
    ]
    if template is not None:
        template_keys = set(template["Page"].map(_normalize_label))
        throughput_data = throughput_data[
            throughput_data["label_key"].isin(template_keys)
        ]

    peak_start = pd.to_datetime(from_date) if from_date is not None else None
    peak_end = pd.to_datetime(to_date) if to_date is not None else None
    if peak_end is not None and peak_end.time() == pd.Timestamp("1900-01-01 00:00:00").time():
        peak_end += pd.Timedelta(days=1)

    formula_start_timestamp = (
        throughput_data["timeStamp"].min() if not throughput_data.empty else None
    )
    formula_end_timestamp = (
        (
            throughput_data["timeStamp"]
            + pd.to_timedelta(throughput_data["elapsed"], unit="ms")
        ).max() if not throughput_data.empty else None
    )

    if throughput_data.empty:
        # Return zero rates when no qualifying transaction remains after filtering.
        duration_sec = 0.0
        received_kb_sec = 0.0
        sent_kb_sec = 0.0
        maximum_throughput_mbps = 0.0
    else:
        # Use the selected peak window when supplied; otherwise use the actual sample time span.
        if peak_start is not None and peak_end is not None:
            duration_sec = (peak_end - peak_start).total_seconds()
        else:
            duration_sec = (formula_end_timestamp - formula_start_timestamp).total_seconds()
        bytes_received = pd.to_numeric(throughput_data["bytes"], errors="coerce").fillna(0).sum()
        bytes_sent = pd.to_numeric(throughput_data["sentBytes"], errors="coerce").fillna(0).sum()
        received_kb_sec = bytes_received / duration_sec / 1024 if duration_sec > 0 else 0.0
        sent_kb_sec = bytes_sent / duration_sec / 1024 if duration_sec > 0 else 0.0
        throughput_data["total_bytes"] = (
            pd.to_numeric(throughput_data["bytes"], errors="coerce").fillna(0)
            + pd.to_numeric(throughput_data["sentBytes"], errors="coerce").fillna(0)
        )
        one_second_totals = throughput_data.groupby(
            throughput_data["timeStamp"].dt.floor("s")
        )["total_bytes"].sum()
        maximum_throughput_mbps = (
            one_second_totals.max() * 8 / 1024 / 1024
            if not one_second_totals.empty else 0.0
        )

    # Convert the combined received and sent KB/sec rates to the final Mbps value.
    avg_throughput = {
        "start_timestamp": peak_start if peak_start is not None else (
            throughput_data["timeStamp"].min() if not throughput_data.empty else None
        ),
        "end_timestamp": peak_end if peak_end is not None else (
            (
                throughput_data["timeStamp"]
                + pd.to_timedelta(throughput_data["elapsed"], unit="ms")
            ).max() if not throughput_data.empty else None
        ),
        "duration_sec": round(duration_sec, 3),
        "received_kb_sec": round(received_kb_sec, 3),
        "sent_kb_sec": round(sent_kb_sec, 3),
        "final_average_throughput_mbps": round(
            (received_kb_sec + sent_kb_sec) * 8 / 1024, 3
        ),
        "maximum_throughput_mbps": round(maximum_throughput_mbps, 3),
    }
    print(
        f"Avg: {avg_throughput['final_average_throughput_mbps']} Mbps, "
        f"Max: {avg_throughput['maximum_throughput_mbps']} Mbps"
    )
    return avg_throughput


calculate_performance_report = calculate_response_jtl


def generate_jtl_summary_sheets(wb, jtl_files, template_file, from_date=None, to_date=None, pass_only=False, keywords=None, data=None):
    """Populate workbook with PassFail and ResponseTime sheets derived from JTL input."""
    template = template_file if isinstance(template_file, pd.DataFrame) else _load_template(template_file)
    data = data if data is not None else load_jtl_data(jtl_files)
    response_summary = calculate_response_jtl(
        jtl_files, from_date, to_date, True, template=template, data=data
    )
    common_page_keys = _common_page_keys(template)
    if common_page_keys:
        common_response_summary = calculate_response_jtl(
            jtl_files, pass_only=True, template=template, data=data
        )
        common_rows = common_response_summary.assign(
            label_key=common_response_summary["label"].map(_normalize_label)
        )
        common_rows = common_rows[common_rows["label_key"].isin(common_page_keys)].set_index("label_key")
        response_keys = response_summary["label"].map(_normalize_label)
        for column in response_summary.columns:
            if column == "label":
                continue
            common_values = response_keys.map(common_rows[column])
            response_summary.loc[response_keys.isin(common_rows.index), column] = common_values[
                response_keys.isin(common_rows.index)
            ]

    # ACTIVE: Use the transaction selection logic below for the PassFail sheet.
    passfail_summary = calculate_response_jtl(
        jtl_files, from_date, to_date, False, template=template, data=data
    )

    # ALTERNATIVE: Use only transactions containing the PassFail keywords instead.
    if response_summary.empty:
        raise ValueError("No JTL data found for the selected time range.")
    if passfail_summary.empty:
        passfail_summary = response_summary.head(0)

    template_lookup = template.set_index("Page")
    passfail_summary = passfail_summary[
        passfail_summary["label"].map(template_lookup["Yes/No"]).eq("yes")
    ].copy()
    passfail_summary = passfail_summary[
        ~passfail_summary["label"].map(_normalize_label).isin(common_page_keys)
    ].copy()
    for field in ("Type", "Module", "Scenario", "Target TPH"):
        response_summary[field] = response_summary["label"].map(template_lookup[field]).fillna("")
        passfail_summary[field] = passfail_summary["label"].map(template_lookup[field]).fillna("")
    standard_template_columns = {
        "Page", "Yes/No", "Type", "Module", "Scenario", "Target TPH"
    }
    extra_passfail_columns = [
        column for column in template.columns
        if column not in standard_template_columns and pd.notna(column)
    ]
    for field in extra_passfail_columns:
        passfail_summary[field] = passfail_summary["label"].map(template_lookup[field]).fillna("")
    template_order = {page: index for index, page in enumerate(template["Page"])}
    response_summary = response_summary.assign(
        __common=response_summary["label"].map(_normalize_label).isin(common_page_keys),
        __template_order=response_summary["label"].map(template_order),
    ).sort_values(["__common", "__template_order"], ascending=[False, True], kind="stable").drop(
        columns=["__common", "__template_order"]
    ).reset_index(drop=True)
    passfail_summary = passfail_summary.assign(
        __template_order=passfail_summary["label"].map(template_order)
    ).sort_values("__template_order", kind="stable").drop(columns="__template_order").reset_index(drop=True)

    response_by_label = response_summary.set_index('label')['Samples']
    passfail_by_label = passfail_summary.set_index('label')['Samples']

    if from_date is not None and to_date is not None:
        start_dt = pd.to_datetime(from_date)
        end_dt = pd.to_datetime(to_date)
        peak_duration_min = max(0.0, (end_dt - start_dt).total_seconds() / 60)
    else:
        peak_duration_min = 1.0

# PassFail sheet: create the report layout and write transaction-level counts and formulas.
#==================================================================================
    passfail_ws = wb.create_sheet(title='PassFail', index=1)
    passfail_ws['B2'] = 'Pass Fail:'
    passfail_ws['B2'].font = Font(name='Cambria', bold=True, underline='single')
    passfail_ws.freeze_panes = f'A{start_row + 1}'
    passfail_ws.sheet_view.showGridLines = False

    passfail_df = pd.DataFrame({
        'Sr. No.': [f'=ROW()-{start_row}' for _ in range(len(passfail_summary))],
        'Type': passfail_summary['Type'].map(lambda v: None if pd.isna(v) or not str(v).strip() else str(v)),
        'Module': passfail_summary['Module'].map(lambda v: None if pd.isna(v) or not str(v).strip() else str(v)),
    })
    for column in extra_passfail_columns:
        passfail_df[column] = passfail_summary[column].map(
            lambda v: None if pd.isna(v) or not str(v).strip() else str(v)
        )
    passfail_df['Scenario'] = passfail_summary['Scenario'].map(
        lambda v: None if pd.isna(v) or not str(v).strip() else str(v)
    )
    passfail_df['Peak Duration in Min'] = peak_duration_min
    passfail_df['Users'] = passfail_summary['Users'].map(
        lambda v: 0 if pd.isna(v) or not str(v).strip() else v
    ) if 'Users' in passfail_summary.columns else 0
    passfail_df['Iterations'] = passfail_summary['Samples'].astype(int)
    passfail_df['Pass'] = response_by_label.reindex(
        passfail_summary['label']
    ).fillna(0).astype(int).to_numpy()
    passfail_df['Fail'] = 0
    passfail_df['Pass%'] = ''
    passfail_df['TPH Achieved'] = ''
    passfail_df['Target TPH'] = passfail_summary['Target TPH'].map(
        lambda v: None if pd.isna(v) or not str(v).strip() else v
    ).astype(object).to_numpy()

    passfail_df = passfail_df[
        ['Sr. No.', 'Type', 'Module', *extra_passfail_columns, 'Scenario',
         'Peak Duration in Min', 'Users', 'Iterations', 'Pass', 'Fail', 'Pass%',
         'TPH Achieved', 'Target TPH']
    ]

    for c_idx, col_name in enumerate(passfail_df.columns, start=start_col):
        cell = passfail_ws.cell(row=start_row, column=c_idx, value=col_name)
        cell.border = thin_border
        cell.fill = Header_fill
        cell.font = white_bold
        cell.alignment = center

    passfail_widths = {
        'Sr. No.': 7,
        'Module': 16,
        'Scenario': 40,
        'Peak Duration in Min': 11,
        'Users': 10,
        'Iterations': 10,
        'TPH Achieved': 10,
        'Target TPH': 10,
    }
    for c_idx, col_name in enumerate(passfail_df.columns, start=start_col):
        if col_name in passfail_widths:
            passfail_ws.column_dimensions[get_column_letter(c_idx)].width = passfail_widths[col_name]

    type_col = passfail_df.columns.get_loc('Type')
    module_col = passfail_df.columns.get_loc('Module')
    duration_col = passfail_df.columns.get_loc('Peak Duration in Min')
    users_col = passfail_df.columns.get_loc('Users')
    iterations_col = passfail_df.columns.get_loc('Iterations')
    pass_col = passfail_df.columns.get_loc('Pass')
    fail_col = passfail_df.columns.get_loc('Fail')
    percent_col = passfail_df.columns.get_loc('Pass%')
    tph_col = passfail_df.columns.get_loc('TPH Achieved')

    type_groups = []
    type_values = passfail_df['Type'].fillna('').astype(str)
    for type_name in type_values.drop_duplicates().tolist():
        type_df = passfail_df[type_values.eq(type_name)].copy()
        if not type_df.empty:
            type_groups.append((type_name, type_df))

    current_row = start_row + 1
    passfail_total_rows = []

    for group_index, (type_name, type_df) in enumerate(type_groups):
        data_start_row = current_row

        for i, row_data in enumerate(type_df.values, start=data_start_row):
            for c_idx, (col_name, val) in enumerate(
                zip(type_df.columns, row_data), start=start_col
            ):
                if col_name == 'Sr. No.':
                    val = f'=ROW()-{start_row + group_index}'
                cell = passfail_ws.cell(
                    row=i,
                    column=c_idx,
                    value=None if pd.isna(val) else val
                )
                cell.border = thin_border
                cell.font = Font(name='Cambria')
                cell.alignment = left if col_name == 'Scenario' else center

            it_col = get_column_letter(start_col + iterations_col)
            p_col = get_column_letter(start_col + pass_col)
            f_col = get_column_letter(start_col + fail_col)
            dur_col = get_column_letter(start_col + duration_col)

            passfail_ws.cell(
                i, start_col + fail_col,
                f'={it_col}{i}-{p_col}{i}'
            )
            passfail_ws.cell(
                i, start_col + percent_col,
                f'=IFERROR({p_col}{i}/{it_col}{i},0)'
            )
            passfail_ws.cell(
                i, start_col + tph_col,
                f'=IFERROR({p_col}{i}*60/{dur_col}{data_start_row},0)'
            )
            passfail_ws.cell(
                i, start_col + percent_col
            ).number_format = '0.00%'

        data_end_row = data_start_row + len(type_df) - 1
        total_row = data_end_row + 1
        passfail_total_rows.append(total_row)

        passfail_ws.cell(
            total_row,
            start_col,
            f'{type_name} TOTAL' if type_name else 'TOTAL'
        )

        users_letter = get_column_letter(start_col + users_col)
        iterations_letter = get_column_letter(start_col + iterations_col)
        pass_letter = get_column_letter(start_col + pass_col)
        fail_letter = get_column_letter(start_col + fail_col)
        duration_letter = get_column_letter(start_col + duration_col)

        target_tph_col = passfail_df.columns.get_loc('Target TPH')
        target_tph_letter = get_column_letter(start_col + target_tph_col)

        passfail_ws.cell(
            total_row,
            start_col + users_col,
            f'=SUM({users_letter}{data_start_row}:{users_letter}{data_end_row})'
        )
        passfail_ws.cell(
            total_row,
            start_col + iterations_col,
            f'=SUM({iterations_letter}{data_start_row}:{iterations_letter}{data_end_row})'
        )
        passfail_ws.cell(
            total_row,
            start_col + pass_col,
            f'=SUM({pass_letter}{data_start_row}:{pass_letter}{data_end_row})'
        )
        passfail_ws.cell(
            total_row,
            start_col + fail_col,
            f'=SUM({fail_letter}{data_start_row}:{fail_letter}{data_end_row})'
        )
        passfail_ws.cell(
            total_row,
            start_col + percent_col,
            f'=IFERROR({pass_letter}{total_row}/{iterations_letter}{total_row},0)'
        )
        passfail_ws.cell(
            total_row,
            start_col + tph_col,
            f'=IFERROR({pass_letter}{total_row}*60/{duration_letter}{data_start_row},0)'
        )

        passfail_ws.cell(
            total_row,
            start_col + target_tph_col,
            f'=SUM({target_tph_letter}{data_start_row}:{target_tph_letter}{data_end_row})'
        )

        passfail_ws.cell(
            total_row,
            start_col + percent_col
        ).number_format = '0.00%'

        for col in range(start_col, start_col + len(passfail_df.columns)):
            cell = passfail_ws.cell(total_row, col)
            cell.border = thin_border
            cell.font = bold
            cell.fill = fill
            cell.alignment = center

        passfail_ws.merge_cells(
            start_row=total_row,
            start_column=start_col,
            end_row=total_row,
            end_column=start_col + 2 + len(extra_passfail_columns)
        )

        passfail_ws.merge_cells(
            start_row=data_start_row,
            start_column=start_col + type_col,
            end_row=data_end_row,
            end_column=start_col + type_col
        )

        for col_name in ['Module', *extra_passfail_columns, 'Scenario']:
            if col_name in type_df.columns:
                _merge_consecutive_cells(
                    passfail_ws,
                    type_df[col_name].tolist(),
                    start_col + type_df.columns.get_loc(col_name),
                    data_start_row
                )

        # _merge_consecutive_cells(passfail_ws, passfail_df['Target TPH'].tolist(), start_col + passfail_df.columns.get_loc('Target TPH'), data_start_row)

        passfail_ws.merge_cells(
            start_row=data_start_row,
            start_column=start_col + duration_col,
            end_row=total_row,
            end_column=start_col + duration_col
        )
        passfail_ws.cell(
            data_start_row,
            start_col + duration_col
        ).alignment = center

        current_row = total_row + 1

    grand_total_row = current_row

    passfail_ws.merge_cells(
        start_row=grand_total_row,
        start_column=start_col,
        end_row=grand_total_row,
        end_column=start_col + 2 + len(extra_passfail_columns)
    )
    passfail_ws.cell(grand_total_row, start_col, 'Total UI + API')

    for col_name in ['Users', 'Iterations', 'Pass', 'Fail']:
        col_index = passfail_df.columns.get_loc(col_name)
        col_letter = get_column_letter(start_col + col_index)
        refs = '+'.join(f'{col_letter}{row}' for row in passfail_total_rows)
        passfail_ws.cell(
            grand_total_row,
            start_col + col_index,
            f'={refs}' if refs else '=0'
        )

    iterations_letter = get_column_letter(start_col + iterations_col)
    pass_letter = get_column_letter(start_col + pass_col)
    duration_letter = get_column_letter(start_col + duration_col)
    target_tph_col = passfail_df.columns.get_loc('Target TPH')
    target_tph_letter = get_column_letter(start_col + target_tph_col)

    passfail_ws.cell(
        grand_total_row,
        start_col + percent_col,
        f'=IFERROR({pass_letter}{grand_total_row}/{iterations_letter}{grand_total_row},0)'
    )
    passfail_ws.cell(
        grand_total_row,
        start_col + tph_col,
        f'=IFERROR({pass_letter}{grand_total_row}*60/{duration_letter}{start_row + 1},0)'
    )

    # passfail_ws.cell(
    #     grand_total_row,
    #     start_col + target_tph_col,
    #     f'=SUM({target_tph_letter}{start_row + 1}:{target_tph_letter}{grand_total_row - 1})'
    # )

    target_tph_refs = '+'.join(
        f'{target_tph_letter}{row}'
        for row in passfail_total_rows
    )

    passfail_ws.cell(
        grand_total_row,
        start_col + target_tph_col,
        f'={target_tph_refs}' if target_tph_refs else '=0'
    )

#==============================================

    passfail_ws.cell(
        grand_total_row,
        start_col + percent_col
    ).number_format = '0.00%'

    for col in range(start_col, start_col + len(passfail_df.columns)):
        cell = passfail_ws.cell(grand_total_row, col)
        cell.border = thin_border
        cell.font = bold
        cell.fill = fill
        cell.alignment = center

#==========================================================================================
# ResponseTime sheet: write successful transaction timing statistics and percentile results.

    response_ws = wb.create_sheet(title='ResponseTime', index=2)
    response_ws.sheet_view.showGridLines = False
    response_ws['I1'] = 'Page response time > 3 sec'
    response_ws['I1'].font = Font(name='Cambria', bold=True)
    response_ws['I1'].fill = highlight_90

    cell = response_ws.cell(row=2, column=3)
    cell.value = 'Page Response Time in seconds: (Only successful transactions)'
    cell.font = Font(name='Cambria', bold=True, underline='single')

    # Preserve common pages first, then the exact template Page order.
    response_summary = response_summary.copy()

    response_df = pd.DataFrame({
        'Type': response_summary['Type'].astype(str).to_numpy(),
        'Module': response_summary['Module'].astype(str).to_numpy(),
        'Scenario': response_summary['Scenario'].astype(str).to_numpy(),
        'Transaction Name': response_summary['label'].astype(str),
        'Minimum': response_summary['Min'].round(2),
        'Average': response_summary['Avg'].round(2),
        'Maximum': response_summary['Max'].round(2),
        '90th %ile': response_summary['P90'].round(2),
        '95th %ile': response_summary['P95'].round(2),
        '99th %ile': response_summary['P99'].round(2),
    })

    response_ws.freeze_panes = f'A{start_row + 1}'

    for c_idx, col_name in enumerate(response_df.columns, start=start_col):
        cell = response_ws.cell(row=start_row, column=c_idx, value=col_name)
        cell.border = thin_border
        cell.fill = Header_fill
        cell.font = white_bold
        cell.alignment = center

    response_widths = {
        'Module': 15,
        'Scenario': 28,
        'Transaction Name': 75,
        'Minimum': 10,
        'Maximum': 10,
    }
    for c_idx, col_name in enumerate(response_df.columns, start=start_col):
        if col_name in response_widths:
            response_ws.column_dimensions[get_column_letter(c_idx)].width = response_widths[col_name]

    for r_idx, row_data in enumerate(response_df.values, start=start_row + 1):
        # Apply consistent borders, alignment, and font styling to each response-time row.
        for c_idx, val in enumerate(row_data, start=start_col):
            cell = response_ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = thin_border
            cell.font = Font(name='Cambria')
            if c_idx == start_col + 3:
                cell.alignment = left
            else:
                cell.alignment = center

    _merge_consecutive_cells(response_ws, response_df['Type'].tolist(), start_col, start_row + 1)
    _merge_consecutive_cells(response_ws, response_df['Module'].tolist(), start_col + 1, start_row + 1)
    _merge_consecutive_cells(response_ws, response_df['Scenario'].tolist(), start_col + 2, start_row + 1)

    for row in range(start_row + 1, start_row + len(response_df) + 1):
        response_ws.cell(row=row, column=start_col + 2).alignment = center

#=====================================================================================

    # Locate the 90th-percentile column so threshold highlighting and counts remain column-independent.
    response_total_row = len(response_df) + start_row
    ninety_col_letter = None
    for idx, cell_obj in enumerate(response_ws[start_row], start=1):
        if cell_obj.value is not None and '90' in str(cell_obj.value).replace('%', '').strip().lower():
            ninety_col_letter = get_column_letter(idx)
            break

    if ninety_col_letter:
        # Highlight response times above three seconds in the 90th-percentile column.
        rule = CellIsRule(operator='greaterThan', formula=['3'], fill=highlight_90, font=Font(color='FF9C0006'))
        response_ws.conditional_formatting.add(f'{ninety_col_letter}{start_row + 1}:{ninety_col_letter}{response_total_row}', rule)

    # Add UI, API, and Total counters for each response-time range shown in the report.
    headers = ['Scenario', 'UI', 'API', 'Total']
    scenario_ranges = [
        ('>=1 & <3', '1', '3'),
        ('>=3 & <5', '3', '5'),
        ('>=5 & <10', '5', '10'),
        ('>=10 & <20', '10', '20'),
        ('>=20 & <30', '20', '30'),
        #======
        ('Total >=1', '1', '50'),
    ]
    for column_offset, header in enumerate(headers):
        header_cell = response_ws.cell(row=1, column=13 + column_offset, value=header)
        header_cell.border = thin_border
        header_cell.font = bold

    for row_offset, (scenario, lower_bound, upper_bound) in enumerate(scenario_ranges, start=2):
        scenario_cell = response_ws.cell(row=row_offset, column=13, value=scenario)
        scenario_cell.border = thin_border
        scenario_cell.font = bold

    if ninety_col_letter:
        ninety_range = f'{ninety_col_letter}{start_row + 1}:{ninety_col_letter}{response_total_row}'
        label_range = f'E{start_row + 1}:E{response_total_row}'

        for row_offset, (_, lower_bound, upper_bound) in enumerate(scenario_ranges, start=2):
            # API means a label starts with API or contains token; overlapping matches count once.
            if row_offset == 7:
                response_ws.cell(row=row_offset, column=14, value='=SUM(N2:N6)')
                response_ws.cell(row=row_offset, column=15, value='=SUM(O2:O6)')
                response_ws.cell(row=row_offset, column=16, value='=SUM(P2:P6)')
                continue

            api_count_formula = (
                f'COUNTIFS({label_range},"API*",{ninety_range},">={lower_bound}",{ninety_range},"<{upper_bound}")+'
                f'COUNTIFS({label_range},"*token*",{ninety_range},">={lower_bound}",{ninety_range},"<{upper_bound}")-'
                f'COUNTIFS({label_range},"API*",{label_range},"*token*",{ninety_range},">={lower_bound}",{ninety_range},"<{upper_bound}")'
            )
            total_count_formula = (
                f'COUNTIFS({ninety_range},">={lower_bound}",{ninety_range},"<{upper_bound}")'
            )

            # UI is the total count for the range minus its API count.
            response_ws.cell(row=row_offset, column=14, value=f'={total_count_formula}-O{row_offset}')
            response_ws.cell(row=row_offset, column=15, value=f'={api_count_formula}')
            response_ws.cell(row=row_offset, column=16, value=f'=N{row_offset}+O{row_offset}')

            #Total Count for UI and API Greater than equal to 1
            # response_ws.cell(row=row_offset, column=15, value=f'={api_count_formula}')
            # response_ws.cell(row=row_offset, column=16, value=f'=N{row_offset}+O{row_offset}')

    for row in range(2, 2 + len(scenario_ranges)):
        for column in range(13, 17):
            response_ws.cell(row=row, column=column).border = thin_border

    # return passfail_total_row, response_total_row
    return passfail_total_rows, grand_total_row, response_total_row

