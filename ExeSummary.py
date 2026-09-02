from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

START_ROW = 4
START_COL = 2

# ARGB colors
dark_header = PatternFill(start_color="FF44546A", end_color="FF44546A", fill_type="solid")
fill = PatternFill(start_color="FFCCCCFF", end_color="FFCCCCFF", fill_type="solid")

white_bold = Font(name="Cambria", color="FFFFFFFF", bold=True)
bold = Font(name="Cambria", bold=True)

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')
)

Cambria_font = Font(name="Cambria")
no_fill = PatternFill(fill_type=None)

# ================= EXECUTIVE SUMMARY =================
def executive_summary(wb, run_id, passfail_total_rows, passfail_total_row, response_total_row, avg_throughput=None):
    pass_ratio_row = None
    ws = wb.create_sheet(title="Executive Summary", index=0)

    def cell(r, c, val="", fill=None, font=None, align=center):
        x = ws.cell(row=r, column=c, value=val)
        x.fill = fill if fill else no_fill
        x.font = font if font else Cambria_font
        x.alignment = align
        return x

    row = 2
    # Title
    ws.merge_cells(start_row=row, start_column=START_COL, end_row=row, end_column=START_COL+4)
    cell(row, 2, "Executive Summary", dark_header, white_bold)

    passfail_ws = wb["PassFail"]

    ui_total_row = passfail_total_rows[0]
    api_total_row = passfail_total_rows[1]

    passfail_headers = {
        passfail_ws.cell(row=START_ROW, column=column).value: column
        for column in range(START_COL, passfail_ws.max_column + 1)
        if passfail_ws.cell(row=START_ROW, column=column).value is not None
    }
    passfail_column = {
        name: get_column_letter(column)
        for name, column in {
            "Scenario": passfail_headers["Scenario"],
            "Users": passfail_headers["Users"],
            "Pass%": passfail_headers["Pass%"],
            "TPH Achieved": passfail_headers["TPH Achieved"],
            "Module": passfail_headers["Module"],
        }.items()
    }

    rows_data1 = [
        ("Run ID", run_id, None),
        ("Test Description", "Load Test was Executed for ", None),
        ("Run Time", "\nPeak Duration:\nTotal Duration:", None),
        ("Scenario Included", f'=CONCAT("UI : ",TEXTJOIN(", ",TRUE,PassFail!{passfail_column["Scenario"]}5:{passfail_column["Scenario"]}{ui_total_row-1}),CHAR(10),"API : ",TEXTJOIN(", ",TRUE,PassFail!{passfail_column["Scenario"]}{ui_total_row+1}:{passfail_column["Scenario"]}{api_total_row}))', None),
        ("Pre-Test Changes", "", None),
        ("Module", f'=TEXTJOIN(", ",TRUE,PassFail!{passfail_column["Module"]}5:{passfail_column["Module"]}{passfail_total_row})', None),
        ("Concurrent Users", f'=CONCAT("UI : ",TEXT(PassFail!{passfail_column["Users"]}{ui_total_row},"0"),",  ","API : ",TEXT(PassFail!{passfail_column["Users"]}{api_total_row},"0"),CHAR(10),"Total : ",TEXT(PassFail!{passfail_column["Users"]}{passfail_total_row},"0"))', None),
        ("TPH", "Achieved", None),
        ("", f'=CONCAT("UI : ",TEXT(PassFail!{passfail_column["TPH Achieved"]}{ui_total_row},"0"),",  ","API : ",TEXT(PassFail!{passfail_column["TPH Achieved"]}{api_total_row},"0"),CHAR(10),"Total : ",TEXT(PassFail!{passfail_column["TPH Achieved"]}{passfail_total_row},"0"))', None),
        ("Pass Ratio %", f'=CONCAT("UI : ",TEXT(PassFail!{passfail_column["Pass%"]}{ui_total_row},"0.00%"),",  ","API : ",TEXT(PassFail!{passfail_column["Pass%"]}{api_total_row},"0.00%"),CHAR(10),"Overall : ",TEXT(PassFail!{passfail_column["Pass%"]}{passfail_total_row},"0.00%"))', None),
        ("Throughput", "" if avg_throughput is None else (
            f"Avg: {avg_throughput['final_average_throughput_mbps']} Mbps, "
            f"Max: {avg_throughput['maximum_throughput_mbps']} Mbps"
        ), None),
        ("", "UI","API"),
        ("Response time @ 90th Percentile >=1 & <3 sec", "=ResponseTime!N2","=ResponseTime!O2"),
        ("Response time @ 90th Percentile >=3 & <5 sec", "=ResponseTime!N3","=ResponseTime!O3"),
        ("Response time @ 90th Percentile >=5 & <10 sec", "=ResponseTime!N4","=ResponseTime!O4"),
        ("Response time @ 90th Percentile >=10 & <20 sec", "=ResponseTime!N5","=ResponseTime!O5"),
        ("Response time @ 90th Percentile >=20 & <30 sec", "=ResponseTime!N6","=ResponseTime!O6"),
        ("", "SLA Status", None),
        ("Pass% >=99.99%", f'=IF(PassFail!{passfail_column["Pass%"]}{passfail_total_row}>=99.99%, "Met", "Not Met")', None),
        ("Page/API Response time @ 90th Percentile < 3 sec", f'=IF(COUNTIF(ResponseTime!I5:I{response_total_row},">3")>0,"Not met","Met")', None),
        ("Infra Utilization <= 70%", "", None),
    ]

    for k, v, w in rows_data1:
        row += 1
        cell(row, START_COL, k, fill, bold)
        if w is not None:
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
            ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
        else:
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)

        # print(v)
        if w is not None:
            cell(row, START_COL+1, v, font=bold)
            cell(row, START_COL+3, w, font=bold)
        elif k in ["Run ID","Module","TPH","Throughput"] or v == "SLA Status":
            cell(row, START_COL+1, v, fill, bold)
        else:
            cell(row, START_COL+1, v)

        if k == "Pass Ratio %":
            pass_ratio_row = row

    ws.merge_cells(start_row=10, start_column=2, end_row=11, end_column=2)

    rows_data2 = [
        ("MicroServices Utilization", "CPU Utilization", "Available Free Memory"),
        ("DB Server", "", ""),
        ("Services", "Refer Pod Services utilization Sheet", "Refer Pod Services utilization Sheet")
    ]

    for k, u, v in rows_data2:
        row += 1
        cell(row, START_COL, k, fill, bold)
        cell(row, START_COL+1, u, fill, bold)
        
        if k == "MicroServices Utilization":
            cell(row, START_COL+1, u, fill, bold)
            cell(row, START_COL+3, v, fill, bold)
        else:
            cell(row, START_COL+1, u)
            cell(row, START_COL+3, v)

        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)


    green_font = Font(name="Cambria", bold=True, color="FF27BB4A")
    red_font = Font(name="Cambria", bold=True, color="FFE20000")

    x = ws.cell(5,3)
    x.font = Font(bold=True)

    ws.conditional_formatting.add(
        "C21:F21",
        CellIsRule(operator='equal', formula=['"Met"'], font=green_font)
    )
    ws.conditional_formatting.add(
        "C21:F21",
        CellIsRule(operator='equal', formula=['"Not Met"'], font=red_font)
    )

    ws.conditional_formatting.add(
        "C22:F22",
        CellIsRule(operator='equal', formula=['"Met"'], font=green_font)
    )
    ws.conditional_formatting.add(
        "C22:F22",
        CellIsRule(operator='equal', formula=['"Not Met"'], font=red_font)
    )

    # ws['C12'].number_format = '0.00%'
    ws.cell(pass_ratio_row, 3).number_format = "0.00%"

    row += 1
    cell(row, START_COL, "Test Conclusion", fill, bold)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    row += 1
    cell(row, START_COL, "")
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)

    ws.column_dimensions['B'].width = 33
    ws.column_dimensions['C'].width = 21
    ws.column_dimensions['D'].width = 21
    ws.column_dimensions['E'].width = 21
    ws.column_dimensions['F'].width = 21
    
    ROW_HEIGHTS = {
        4: 40,
        5: 45,
        9: 30,
        11: 30,
        12: 30,
        13: 17
    }

    for row_number, height in ROW_HEIGHTS.items():
        ws.row_dimensions[row_number].height = height

    # Apply the complete Executive Summary table border in one place after all rows and merges exist.
    for table_row in range(2, row + 1):
        for table_column in range(2, 7):
            ws.cell(row=table_row, column=table_column).border = thin_border

    ws.sheet_view.showGridLines = False